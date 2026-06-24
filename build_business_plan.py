#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ---------- Palette ----------
NAVY = "1F3864"
BLUE = "2E5496"
MIDBLUE = "305496"
LIGHT = "DDEBF7"
GREYBG = "F2F2F2"
GREYLN = "BFBFBF"
KPIBG = "1F3864"
FONT = "Arial"

NOTE_FONT = Font(name=FONT, color="595959", size=9, italic=True)
INPUT_FILL = PatternFill("solid", fgColor=LIGHT)
thin = Side(style="thin", color=GREYLN)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

GNF = '#,##0" GNF"'
USD = '#,##0.0" $"'
USD2 = '#,##0.00" $"'
PCT = '0.0%'
NUM = '#,##0'

wb = Workbook()

def title_bar(ws, text, ncol=4, sub=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(1, 1, text)
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=16)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
        s = ws.cell(2, 1, sub)
        s.font = Font(name=FONT, color="FFFFFF", size=10, italic=True)
        s.fill = PatternFill("solid", fgColor=BLUE)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18

def section(ws, row, text, ncol=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    c = ws.cell(row, 1, text)
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20

def header_row(ws, row, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=MIDBLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 26

def put(ws, row, col, val, *, fmt=None, inp=False, bold=False, note=False,
        align=None, border=True, wrap=False, fill=None, color=None):
    c = ws.cell(row, col, val)
    if note:
        c.font = NOTE_FONT
    else:
        c.font = Font(name=FONT, bold=bold, size=11,
                      color=(color if color else ("0000FF" if inp else "000000")))
    if inp:
        c.fill = INPUT_FILL
    elif fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    else:
        c.alignment = Alignment(vertical="center", wrap_text=wrap)
    if border:
        c.border = BORDER
    return c

# =====================================================================
# 1. PAGE DE GARDE
# =====================================================================
ws = wb.active
ws.title = "Page de garde"
ws.sheet_view.showGridLines = False
for col, w in [("A", 3), ("B", 26), ("C", 40), ("D", 24), ("E", 3)]:
    ws.column_dimensions[col].width = w
ws.merge_cells("B2:D3")
c = ws.cell(2, 2, "HPSHOP")
c.font = Font(name=FONT, bold=True, color=NAVY, size=40)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells("B4:D4")
c = ws.cell(4, 2, "Helpmeprocess — Commerce en ligne (COD) en Guinée")
c.font = Font(name=FONT, color=BLUE, size=14, italic=True)
c.alignment = Alignment(horizontal="center")
ws.merge_cells("B6:D6")
c = ws.cell(6, 2, "BUSINESS PLAN")
c.font = Font(name=FONT, bold=True, color="FFFFFF", size=18)
c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[6].height = 30
ws.merge_cells("B7:D7")
c = ws.cell(7, 2, "Dossier de recherche de financement — Investisseurs & bailleurs de fonds")
c.font = Font(name=FONT, color="FFFFFF", size=11)
c.fill = PatternFill("solid", fgColor=BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[7].height = 22
rows = [
    ("Projet", "HPSHOP — Site e-commerce (Helpmeprocess)"),
    ("Porteur de projet", "Abdourahamane Cissé — XPROCESS"),
    ("Date", "Juin 2026"),
    ("Objet", "Recherche de financement pour le lancement et la structuration"),
    ("Devises", "Dollar US (USD) et Franc guinéen (GNF)"),
]
r = 9
for k, v in rows:
    c = ws.cell(r, 2, k)
    c.font = Font(name=FONT, bold=True, color=NAVY, size=11)
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    c2 = ws.cell(r, 3, v)
    c2.font = Font(name=FONT, size=11)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
c = ws.cell(r, 2, "MODÈLE ÉCONOMIQUE EN UNE PHRASE")
c.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
c.fill = PatternFill("solid", fgColor=BLUE)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[r].height = 20
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=4)
c = ws.cell(r, 2,
    "Achat de produits tendance en Chine / Dubaï via AfriShop, importés par fret jusqu'en Guinée, "
    "revente en ligne sur HPSHOP (paiement à la livraison — COD), avec confirmation des commandes "
    "et livraison sous-traitées à Rapido sur Conakry et l'intérieur du pays.")
c.font = Font(name=FONT, size=12)
c.fill = PatternFill("solid", fgColor=GREYBG)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
for rr in range(r, r + 3):
    ws.row_dimensions[rr].height = 22

# =====================================================================
# 2. POURQUOI INVESTIR
# =====================================================================
ws = wb.create_sheet("Pourquoi investir")
ws.sheet_view.showGridLines = False
for col, w in [("A", 3), ("B", 30), ("C", 30), ("D", 30), ("E", 30), ("F", 3)]:
    ws.column_dimensions[col].width = w
title_bar(ws, "POURQUOI INVESTIR DANS CE MODÈLE", ncol=5,
          sub="Argumentaire stratégique — product-market fit déjà démontré")

def para(ws, row, head, body, ncol=5):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncol)
    c = ws.cell(row, 2, head)
    c.font = Font(name=FONT, bold=True, color=NAVY, size=12)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 2, end_column=ncol)
    c = ws.cell(row + 1, 2, body)
    c.font = Font(name=FONT, size=11)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row + 1].height = 30
    ws.row_dimensions[row + 2].height = 30
    return row + 4

r = 4
r = para(ws, r, "1.  Le COD (paiement à la livraison) — taillé pour le marché guinéen",
    "En Guinée, la bancarisation reste faible et la méfiance envers le paiement en ligne anticipé est forte. "
    "Le COD lève ce frein : le client ne paie qu'à la réception physique du produit. La confiance est instaurée "
    "par l'acte d'achat lui-même, ce qui débloque la conversion là où le prépaiement échoue. C'est le mode de "
    "paiement le plus adapté au contexte local et le socle de notre modèle.")
r = para(ws, r, "2.  L'intérêt du e-commerce en Guinée aujourd'hui",
    "Marché jeune et forte pénétration mobile : Facebook et Instagram servent de vitrine commerciale directe. "
    "Il n'existe pas encore d'acteur structuré sur nos catégories (auto/entretien, beauté, tech accessible, maison). "
    "L'opportunité : devenir un acteur de référence avant l'arrivée de concurrents plus importants et mieux capitalisés.")

section(ws, r, "PREUVE DE TRACTION — CHIFFRES DÉJÀ RÉALISÉS AVEC AFRISHOP", ncol=5)
r += 1
kpis = [
    ("100%", "Taux de livraison\nhistorique"),
    ("45", "Engine Protectant\n(unités vendues)"),
    ("30", "Car Carbon Fiber\nFilm 5D (unités)"),
    ("29", "Crème Visage Kojic\n(unités)"),
]
for j, (big, lab) in enumerate(kpis, 2):
    c = ws.cell(r, j, big)
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=28)
    c.fill = PatternFill("solid", fgColor=KPIBG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER
    c2 = ws.cell(r + 1, j, lab)
    c2.font = Font(name=FONT, color="FFFFFF", size=10)
    c2.fill = PatternFill("solid", fgColor=BLUE)
    c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c2.border = BORDER
ws.row_dimensions[r].height = 46
ws.row_dimensions[r + 1].height = 32
r += 3
bullets = [
    "Top produits déjà validés par la demande réelle : Engine Protectant (45), Car Carbon Fiber Film 5D (30), "
    "Crème Visage Kojic (29), Shilajit Soft Candy et Googles (26 chacun).",
    "Couverture géographique déjà opérationnelle : Conakry (toutes communes) + intérieur du pays "
    "(Coyah, Kamsar, Boké, etc.).",
    "Charges opérationnelles actuelles ≈ 61% du chiffre d'affaires — mentionné en toute transparence. "
    "C'est précisément ce que le nouveau modèle (site propre + catalogue optimisé + Rapido) vise à améliorer.",
]
for b in bullets:
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=5)
    c = ws.cell(r, 2, "•  " + b)
    c.font = Font(name=FONT, size=11)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 18
    ws.row_dimensions[r + 1].height = 18
    r += 2
r += 1
section(ws, r, "CONCLUSION", ncol=5)
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=5)
c = ws.cell(r, 2,
    "Le product-market fit est déjà démontré sur des produits similaires et la demande est confirmée. "
    "Ce business plan présente l'étape de structuration — site propre HPSHOP, catalogue élargi, partenariat "
    "logistique formalisé avec Rapido — pour passer d'un test informel à une activité e-commerce structurée, "
    "pilotable et finançable.")
c.font = Font(name=FONT, size=12, bold=True, color=NAVY)
c.fill = PatternFill("solid", fgColor=GREYBG)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
for rr in range(r, r + 3):
    ws.row_dimensions[rr].height = 24

# =====================================================================
# 3. HYPOTHESES  (USD + GNF uniquement, avec fret)
# =====================================================================
ws = wb.create_sheet("Hypotheses")
ws.sheet_view.showGridLines = False
for col, w in [("A", 44), ("B", 16), ("C", 12), ("D", 46)]:
    ws.column_dimensions[col].width = w
title_bar(ws, "HYPOTHÈSES & TAUX DE CHANGE",
          sub="Cellules en BLEU = modifiables. Devises : USD et GNF. Les calculs (noir) se mettent à jour automatiquement.")
header_row(ws, 3, ["Paramètre", "Valeur", "Unité", "Source / Note"])

def hrow(r, label, val, unit, src, fmt=None, inp=False, calc=False):
    put(ws, r, 1, label, align="left")
    put(ws, r, 2, val, fmt=fmt, inp=inp, align="right", bold=calc,
        color=("000000" if calc else None))
    put(ws, r, 3, unit, align="center")
    put(ws, r, 4, src, note=True, align="left")

section(ws, 4, "TAUX DE CHANGE")
hrow(5, "1 USD en GNF", 8775, "GNF", "Taux moyen marché, juin 2026", fmt=NUM, inp=True)

section(ws, 6, "COÛTS & PRIX PRODUITS")
hrow(7, "Coût d'achat produit — bas (Chine)", 4.0, "USD", "Hypothèse", fmt=USD2, inp=True)
hrow(8, "Coût d'achat produit — haut (Chine)", 6.0, "USD", "Hypothèse", fmt=USD2, inp=True)
hrow(9, "Coût d'achat produit retenu", "=(B7+B8)/2", "USD", "Calcul (milieu)", fmt=USD2, calc=True)
hrow(10, "Coût d'achat produit retenu", "=B9*B5", "GNF", "Calcul (× taux USD)", fmt=GNF, calc=True)
hrow(11, "Fret / transport par article", 2.65, "USD", "Devis transitaire — converti", fmt=USD2, inp=True)
hrow(12, "Fret / transport par article", "=B11*B5", "GNF", "Calcul (× taux USD)", fmt=GNF, calc=True)
hrow(13, "Coût RENDU Guinée / article", "=B9+B11", "USD", "Produit + fret", fmt=USD2, calc=True)
hrow(14, "Coût RENDU Guinée / article", "=B10+B12", "GNF", "Produit + fret", fmt=GNF, calc=True)
hrow(15, "Prix de vente — bas", 180000, "GNF", "Prix marché Guinée, juin 2026", fmt=GNF, inp=True)
hrow(16, "Prix de vente — haut", 250000, "GNF", "Prix marché Guinée, juin 2026", fmt=GNF, inp=True)
hrow(17, "Prix de vente retenu (milieu)", "=(B15+B16)/2", "GNF", "Calcul", fmt=GNF, calc=True)
hrow(18, "Prix de vente retenu — base des calculs", "=B17", "GNF", "Calcul", fmt=GNF, calc=True)

section(ws, 19, "LOGISTIQUE & ACQUISITION")
hrow(20, "Coût Rapido par commande livrée", 70000, "GNF", "Valeur fixe — porteur de projet", fmt=GNF, inp=True)
hrow(21, "Budget pub Meta — bas", 20, "USD/jour", "Révisé (budget renforcé)", fmt=USD, inp=True)
hrow(22, "Budget pub Meta — haut", 40, "USD/jour", "Révisé (budget renforcé)", fmt=USD, inp=True)
hrow(23, "Budget pub Meta retenu (milieu)", "=(B21+B22)/2", "USD/jour", "Calcul", fmt=USD, calc=True)
hrow(24, "Budget pub Meta retenu", "=B23*B5", "GNF/jour", "Calcul (× taux USD)", fmt=GNF, calc=True)

section(ws, 25, "VOLUME & PERFORMANCE")
hrow(26, "Objectif commandes confirmées / jour", 30, "cmd/jour", "Hypothèse modifiable", fmt=NUM, inp=True)
hrow(27, "Taux de livraison réussie (COD)", 0.70, "%",
     "Ajustable — toutes les commandes confirmées ne sont pas livrées", fmt=PCT, inp=True)
hrow(28, "Commandes livrées / jour", "=B26*B27", "cmd/jour", "Calcul", fmt='#,##0.0', calc=True)
hrow(29, "Jours opérés / mois", 30, "jours", "Hypothèse", fmt=NUM, inp=True)

section(ws, 30, "MONTÉE EN CHARGE (RAMP-UP) — % de l'objectif atteint")
ramp = [0.50, 0.75, 1.00, 1.00, 1.00, 1.00]
for i, v in enumerate(ramp):
    hrow(31 + i, f"Mois {i+1}", v, "%",
         "Hypothèse de démarrage progressif" if i == 0 else "", fmt=PCT, inp=True)

ws.cell(5, 2).comment = Comment("Modifier ce taux met à jour toutes les conversions USD→GNF du classeur.", "HPSHOP")
ws.cell(11, 2).comment = Comment(
    "Source : devis transitaire ≈ 1 500 FCFA/article (juin 2026), converti en USD/GNF. "
    "À remplacer par le tarif réel au poids/volume dès qu'il est confirmé.", "HPSHOP")

# ----- Table de références vers Hypotheses -----
def H(addr):  # addr like 'B14'
    col, row = addr[0], addr[1:]
    return f"Hypotheses!${col}${row}"
R = {
    'usd': 'B5', 'cost_usd': 'B9', 'cost_gnf': 'B10', 'fret_usd': 'B11', 'fret_gnf': 'B12',
    'landed_usd': 'B13', 'landed_gnf': 'B14', 'price_gnf': 'B18',
    'rapido': 'B20', 'pub_usd': 'B23', 'pub_gnf': 'B24',
    'obj': 'B26', 'taux': 'B27', 'livr': 'B28', 'jours': 'B29',
}
ramp_rows = [31, 32, 33, 34, 35, 36]

# =====================================================================
# 4. CATALOGUE PRODUITS
# =====================================================================
ws = wb.create_sheet("Catalogue")
ws.sheet_view.showGridLines = False
for i, w in enumerate([40, 26, 16, 18, 16, 16, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
title_bar(ws, "CATALOGUE PRODUITS", ncol=7,
          sub="Coût d'achat (USD) et prix de vente (GNF) en BLEU = modifiables. Coût rendu = achat + fret. Marges calculées.")
header_row(ws, 3, ["Produit", "Catégorie", "Coût achat\n(USD)", "Coût rendu Guinée\n(GNF, fret inclus)",
                   "Prix vente\n(GNF)", "Marge brute\n(GNF)", "Marge\n%"])
products = [
    ("Engine Protectant", "AUTO / ENTRETIEN"),
    ("Car Carbon Fiber Film 5D Black", "AUTO / ENTRETIEN"),
    ("Crème d'entretien pour cuir", "AUTO / ENTRETIEN"),
    ("Crème Visage Kojic", "BEAUTÉ & BIEN-ÊTRE"),
    ("Shilajit Soft Candy", "BEAUTÉ & BIEN-ÊTRE"),
    ("Tooth Whitening Powder", "BEAUTÉ & BIEN-ÊTRE"),
    ("Electric Nail Clippers", "BEAUTÉ & BIEN-ÊTRE"),
    ("Masques de protection moto (Goggles)", "TECH & JEUNESSE"),
    ("FlexiMag 3-in-1", "TECH & JEUNESSE"),
    ("Chargeur de secours / power bank", "TECH & JEUNESSE"),
    ("Pack Smartwatch + écouteurs (iRed Ultra)", "TECH & JEUNESSE"),
    ("Caméra surveillance IP ampoule 360°", "ÉLECTROMÉNAGER & MAISON"),
    ("Ventilateur brumisateur double tête", "ÉLECTROMÉNAGER & MAISON"),
    ("Kit santé (tensiomètre+glycémie+oxymètre)", "ÉLECTROMÉNAGER & MAISON"),
]
r = 4
for name, cat in products:
    put(ws, r, 1, name, align="left")
    put(ws, r, 2, cat, align="left")
    put(ws, r, 3, 5.0, fmt=USD2, inp=True, align="right")  # coût achat USD (défaut milieu)
    put(ws, r, 4, f"=C{r}*{H(R['usd'])}+{H(R['fret_gnf'])}", fmt=GNF, align="right")  # rendu = achat*taux + fret
    put(ws, r, 5, 215000, fmt=GNF, inp=True, align="right")  # prix vente GNF (défaut milieu)
    put(ws, r, 6, f"=E{r}-D{r}", fmt=GNF, align="right")
    put(ws, r, 7, f"=F{r}/E{r}", fmt=PCT, align="right")
    r += 1
put(ws, r, 1, "MOYENNE CATALOGUE", bold=True, align="left", fill=GREYBG)
put(ws, r, 2, "", fill=GREYBG)
put(ws, r, 3, f"=AVERAGE(C4:C{r-1})", fmt=USD2, bold=True, align="right", fill=GREYBG)
put(ws, r, 4, f"=AVERAGE(D4:D{r-1})", fmt=GNF, bold=True, align="right", fill=GREYBG)
put(ws, r, 5, f"=AVERAGE(E4:E{r-1})", fmt=GNF, bold=True, align="right", fill=GREYBG)
put(ws, r, 6, f"=AVERAGE(F4:F{r-1})", fmt=GNF, bold=True, align="right", fill=GREYBG)
put(ws, r, 7, f"=F{r}/E{r}", fmt=PCT, bold=True, align="right", fill=GREYBG)

# =====================================================================
# 5. COUTS PAR COMMANDE
# =====================================================================
ws = wb.create_sheet("Couts par commande")
ws.sheet_view.showGridLines = False
for col, w in [("A", 50), ("B", 18), ("C", 44)]:
    ws.column_dimensions[col].width = w
title_bar(ws, "COÛTS PAR COMMANDE — RENTABILITÉ UNITAIRE", ncol=3,
          sub="Le fret/transport apparaît comme une ligne de coût distincte. Valeurs issues de l'onglet Hypothèses.")
section(ws, 3, "A. POUR UNE COMMANDE LIVRÉE", ncol=3)
header_row(ws, 4, ["Poste", "Montant", "Note"])
rows = [
    ("Prix de vente (moyen)", f"={H(R['price_gnf'])}", GNF, "Catalogue — prix retenu"),
    ("(–) Coût d'achat produit", f"={H(R['cost_gnf'])}", GNF, "Coût Chine converti en GNF"),
    ("(–) Fret / transport par article", f"={H(R['fret_gnf'])}", GNF, "Devis transitaire (≈ 2,65 $)"),
    ("(–) Coût Rapido (confirmation + livraison)", f"={H(R['rapido'])}", GNF, "70 000 GNF — valeur fixe"),
    ("(–) Coût publicitaire alloué / commande livrée", f"={H(R['pub_gnf'])}/{H(R['livr'])}", GNF,
     "Budget pub jour ÷ commandes livrées/jour"),
    ("►  Marge nette par commande livrée", "=B5-B6-B7-B8-B9", GNF, "Calcul"),
    ("Marge nette (%)", "=B10/B5", PCT, "Calcul"),
]
r = 5
for lab, val, fmt, note in rows:
    bold = lab.startswith("►")
    put(ws, r, 1, lab, align="left", bold=bold, fill=(GREYBG if bold else None))
    put(ws, r, 2, val, fmt=fmt, align="right", bold=bold, fill=(GREYBG if bold else None))
    put(ws, r, 3, note, note=True, align="left", fill=(GREYBG if bold else None))
    r += 1
r += 1
section(ws, r, "B. IMPACT DU TAUX DE LIVRAISON (par jour)", ncol=3); r += 1
header_row(ws, r, ["Poste", "Montant", "Note"]); r += 1
d = r
drows = [
    ("Commandes confirmées / jour", f"={H(R['obj'])}", NUM, ""),
    ("Taux de livraison réussie", f"={H(R['taux'])}", PCT, ""),
    ("Commandes livrées / jour", f"={H(R['livr'])}", '#,##0.0', ""),
    ("Commandes confirmées non livrées / jour", f"=B{d}-B{d+2}", '#,##0.0', "Perte sèche pub"),
    ("Chiffre d'affaires / jour", f"=B{d+2}*B5", GNF, "Livrées × prix"),
    ("(–) Coût des produits / jour", f"=B{d+2}*B6", GNF, ""),
    ("(–) Coût du fret / jour", f"=B{d+2}*B7", GNF, "Transport des articles livrés"),
    ("(–) Coût Rapido / jour", f"=B{d+2}*B8", GNF, ""),
    ("(–) Coût publicitaire / jour", f"={H(R['pub_gnf'])}", GNF, "Dépensé quel que soit le taux"),
    ("►  Marge nette / jour", f"=B{d+4}-B{d+5}-B{d+6}-B{d+7}-B{d+8}", GNF, "Calcul"),
    ("Perte pub sur commandes non livrées / jour", f"=B{d+3}*({H(R['pub_gnf'])}/{H(R['obj'])})", GNF,
     "Budget pub gaspillé sur non-livrées"),
]
for lab, val, fmt, note in drows:
    bold = lab.startswith("►")
    put(ws, r, 1, lab, align="left", bold=bold, fill=(GREYBG if bold else None))
    put(ws, r, 2, val, fmt=fmt, align="right", bold=bold, fill=(GREYBG if bold else None))
    put(ws, r, 3, note, note=True, align="left", fill=(GREYBG if bold else None))
    r += 1
couts_marge_nette_cell = "'Couts par commande'!$B$10"

# =====================================================================
# 6. BUDGET PUB META
# =====================================================================
ws = wb.create_sheet("Budget Pub Meta")
ws.sheet_view.showGridLines = False
for col, w in [("A", 42), ("B", 18), ("C", 16), ("D", 18), ("E", 18)]:
    ws.column_dimensions[col].width = w
title_bar(ws, "BUDGET PUBLICITÉ META (Facebook / Instagram)", ncol=5,
          sub="Budget renforcé : 20 à 40 USD/jour (auparavant 5–10).")
header_row(ws, 3, ["Paramètre", "Valeur", "Unité", "", ""])
b = [
    ("Budget pub / jour (retenu)", f"={H(R['pub_usd'])}", USD, "USD/jour"),
    ("Budget pub / jour (retenu)", f"={H(R['pub_gnf'])}", GNF, "GNF/jour"),
    ("Jours / mois", f"={H(R['jours'])}", NUM, "jours"),
    ("Budget pub / mois", "=B4*B6", USD, "USD/mois"),
    ("Budget pub / mois", "=B5*B6", GNF, "GNF/mois"),
    ("Commandes confirmées visées / jour", f"={H(R['obj'])}", NUM, "cmd/jour"),
    ("Commandes confirmées visées / mois", "=B9*B6", NUM, "cmd/mois"),
    ("Commandes livrées / mois", f"={H(R['livr'])}*B6", '#,##0.0', "cmd/mois"),
    ("Coût d'acquisition / commande confirmée", "=B5/B9", GNF, "CPA confirmée"),
    ("Coût d'acquisition / commande livrée", f"=B5/{H(R['livr'])}", GNF, "CPA livrée"),
]
r = 4
for lab, val, fmt, unit in b:
    put(ws, r, 1, lab, align="left")
    put(ws, r, 2, val, fmt=fmt, align="right")
    put(ws, r, 3, unit, align="center")
    r += 1
r += 1
section(ws, r, "SCÉNARIOS DE BUDGET", ncol=5); r += 1
header_row(ws, r, ["Scénario", "USD / jour", "GNF / jour", "GNF / mois", "CPA confirmée"]); r += 1
sc = [("Prudent", 20), ("Retenu", f"={H(R['pub_usd'])}"), ("Agressif", 40)]
for name, usd in sc:
    put(ws, r, 1, name, align="left")
    put(ws, r, 2, usd, fmt=USD, align="right", inp=isinstance(usd, (int, float)))
    put(ws, r, 3, f"=B{r}*{H(R['usd'])}", fmt=GNF, align="right")
    put(ws, r, 4, f"=C{r}*{H(R['jours'])}", fmt=GNF, align="right")
    put(ws, r, 5, f"=C{r}/{H(R['obj'])}", fmt=GNF, align="right")
    r += 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=5)
c = ws.cell(r, 1,
    "Note : ces hypothèses s'appuient sur des tests publicitaires réels déjà menés par l'entreprise "
    "sur le marché guinéen — et non sur des moyennes Facebook Europe/USA, non pertinentes ici.")
c.font = NOTE_FONT
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# =====================================================================
# 7. MISE EN PLACE SITE (CAPEX)
# =====================================================================
ws = wb.create_sheet("Mise en place site")
ws.sheet_view.showGridLines = False
for col, w in [("A", 46), ("B", 18), ("C", 40)]:
    ws.column_dimensions[col].width = w
title_bar(ws, "MISE EN PLACE DU SITE HPSHOP — CAPEX", ncol=3,
          sub="Montants en BLEU = estimations à valider avant envoi au bailleur.")
header_row(ws, 3, ["Poste", "Montant (GNF)", "Note / Statut"])
capex = [
    ("Développement & personnalisation du site HPSHOP", 5000000, "Estimation à valider"),
    ("Design, branding & logo", 1500000, "Estimation à valider"),
    ("Intégration paiement / COD / formulaire de commande", 1000000, "Estimation à valider"),
    ("Photos & contenu produits", 1000000, "Estimation à valider"),
    ("Nom de domaine (1 an)", 300000, "Estimation à valider"),
    ("Hébergement (1 an)", 1200000, "Estimation à valider"),
    ("Frais ponctuels divers (configuration, tests)", 500000, "Estimation à valider"),
]
r = 4
for lab, val, note in capex:
    put(ws, r, 1, lab, align="left")
    put(ws, r, 2, val, fmt=GNF, inp=True, align="right")
    put(ws, r, 3, note, note=True, align="left")
    r += 1
put(ws, r, 1, "TOTAL CAPEX SITE", bold=True, align="left", fill=NAVY, color="FFFFFF")
put(ws, r, 2, f"=SUM(B4:B{r-1})", fmt=GNF, bold=True, align="right", fill=NAVY, color="FFFFFF")
put(ws, r, 3, "", fill=NAVY)
capex_total_row = r
r += 2
section(ws, r, "COÛTS FIXES MENSUELS RÉCURRENTS", ncol=3); r += 1
header_row(ws, r, ["Poste", "Montant (GNF)", "Note"]); r += 1
put(ws, r, 1, "Maintenance & mises à jour du site / mois", align="left")
put(ws, r, 2, 500000, fmt=GNF, inp=True, align="right")
put(ws, r, 3, "Estimation à valider", note=True, align="left")
m1 = r; r += 1
put(ws, r, 1, "Hébergement + domaine amortis / mois", align="left")
put(ws, r, 2, "=(B8+B9)/12", fmt=GNF, align="right")
put(ws, r, 3, "Calcul (postes annuels ÷ 12)", note=True, align="left")
m2 = r; r += 1
put(ws, r, 1, "TOTAL COÛT FIXE MENSUEL", bold=True, align="left", fill=NAVY, color="FFFFFF")
put(ws, r, 2, f"=SUM(B{m1}:B{m2})", fmt=GNF, bold=True, align="right", fill=NAVY, color="FFFFFF")
put(ws, r, 3, "", fill=NAVY)
fixe_mensuel_row = r

# =====================================================================
# 8. APPROVISIONNEMENT
# =====================================================================
ws = wb.create_sheet("Approvisionnement")
ws.sheet_view.showGridLines = False
for col, w in [("A", 46), ("B", 18), ("C", 12), ("D", 40)]:
    ws.column_dimensions[col].width = w
title_bar(ws, "APPROVISIONNEMENT / SUPPLY CHAIN",
          sub="Cellules en BLEU = à compléter / ajuster (délais, coefficient). Fret intégré depuis l'onglet Hypothèses.")
header_row(ws, 3, ["Paramètre", "Valeur", "Unité", "Note"])
def arow(r, lab, val, unit, note, fmt=None, inp=False):
    put(ws, r, 1, lab, align="left")
    put(ws, r, 2, val, fmt=fmt, inp=inp, align="right")
    put(ws, r, 3, unit, align="center")
    put(ws, r, 4, note, note=True, align="left")
arow(4, "Coût d'achat produit (Chine)", f"={H(R['cost_usd'])}", "USD", "Hypothèses", fmt=USD2)
arow(5, "Coût d'achat produit (Chine)", f"={H(R['cost_gnf'])}", "GNF", "Hypothèses", fmt=GNF)
arow(6, "Fret / transport par article", f"={H(R['fret_usd'])}", "USD", "Devis transitaire — converti en USD/GNF", fmt=USD2)
arow(7, "Fret / transport par article", f"={H(R['fret_gnf'])}", "GNF", "Devis transitaire", fmt=GNF)
arow(8, "Coût RENDU Guinée / article", f"={H(R['landed_gnf'])}", "GNF", "Produit + fret", fmt=GNF)
section(ws, 9, "VOLUME & STOCK", ncol=4)
arow(10, "Commandes confirmées / jour (plein régime)", f"={H(R['obj'])}", "cmd/j", "Hypothèses", fmt=NUM)
arow(11, "Délai d'approvisionnement estimé", 30, "jours", "ESTIMATION — à compléter", fmt=NUM, inp=True)
arow(12, "Coefficient de stock de sécurité", 1.5, "x", "ESTIMATION — marge de sécurité", fmt='#,##0.0', inp=True)
arow(13, "Stock de sécurité recommandé", "=B10*B11*B12", "unités", "Calcul (cmd/j × délai × coef)", fmt=NUM)
arow(14, "Volume mensuel livré (plein régime)", f"={H(R['livr'])}*{H(R['jours'])}", "unités", "Calcul", fmt=NUM)
arow(15, "Coût d'achat mensuel (rendu Guinée)", f"=B14*{H(R['landed_gnf'])}", "GNF", "Calcul", fmt=GNF)
arow(16, "Valeur du stock de sécurité (stock initial)", f"=B13*{H(R['landed_gnf'])}", "GNF",
     "Calcul — sert au plan de financement", fmt=GNF)
ws.cell(11, 2).comment = Comment("Délai porte-à-porte estimé Chine→Conakry. À confirmer avec le transitaire.", "HPSHOP")
appro_stock_row = 16

# =====================================================================
# 9. RESULTAT 6 MOIS (P&L)
# =====================================================================
ws = wb.create_sheet("Resultat 6 mois")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 40
for i in range(2, 9):
    ws.column_dimensions[get_column_letter(i)].width = 15
title_bar(ws, "COMPTE DE RÉSULTAT PRÉVISIONNEL — 6 MOIS", ncol=8,
          sub="Montée en charge progressive (ramp-up) modifiable dans l'onglet Hypothèses. Montants en GNF.")
header_row(ws, 3, ["Poste (GNF)", "Mois 1", "Mois 2", "Mois 3", "Mois 4", "Mois 5", "Mois 6", "Total 6 mois"])
mcols = [2, 3, 4, 5, 6, 7]
def L(i): return get_column_letter(i)

put(ws, 4, 1, "% de l'objectif (ramp-up)", align="left", fill=GREYBG, bold=True)
for k, ci in enumerate(mcols):
    put(ws, 4, ci, f"=Hypotheses!$B${ramp_rows[k]}", fmt=PCT, align="right", fill=GREYBG)
put(ws, 4, 8, "", fill=GREYBG)
put(ws, 5, 1, "Commandes confirmées / jour", align="left")
for ci in mcols:
    put(ws, 5, ci, f"={H(R['obj'])}*{L(ci)}4", fmt='#,##0.0', align="right")
put(ws, 5, 8, "", border=True)
put(ws, 6, 1, "Commandes confirmées / mois", align="left")
for ci in mcols:
    put(ws, 6, ci, f"={L(ci)}5*{H(R['jours'])}", fmt=NUM, align="right")
put(ws, 6, 8, "=SUM(B6:G6)", fmt=NUM, align="right", bold=True)
put(ws, 7, 1, "Commandes livrées / mois", align="left")
for ci in mcols:
    put(ws, 7, ci, f"={L(ci)}6*{H(R['taux'])}", fmt=NUM, align="right")
put(ws, 7, 8, "=SUM(B7:G7)", fmt=NUM, align="right", bold=True)

def pnl(r, lab, tpl, total=True, bold=False, fill=None, color=None):
    put(ws, r, 1, lab, align="left", bold=bold, fill=fill, color=color)
    for ci in mcols:
        put(ws, r, ci, tpl.format(L=L(ci)), fmt=GNF, align="right", bold=bold, fill=fill, color=color)
    if total:
        put(ws, r, 8, f"=SUM(B{r}:G{r})", fmt=GNF, align="right", bold=True, fill=fill, color=color)
    else:
        put(ws, r, 8, "", fill=fill)

pnl(9,  "Chiffre d'affaires", "={L}7*" + H(R['price_gnf']))
pnl(10, "(–) Coût des produits vendus", "={L}7*" + H(R['cost_gnf']))
pnl(11, "(–) Coût du fret / transport", "={L}7*" + H(R['fret_gnf']))
pnl(12, "►  Marge brute", "={L}9-{L}10-{L}11", bold=True, fill=GREYBG)
pnl(13, "(–) Coûts Rapido", "={L}7*" + H(R['rapido']))
pnl(14, "(–) Coûts publicitaires", "=" + H(R['pub_gnf']) + "*" + H(R['jours']))
pnl(15, "(–) Coûts fixes (site)", f"='Mise en place site'!$B${fixe_mensuel_row}")
pnl(16, "►  Résultat net", "={L}12-{L}13-{L}14-{L}15", bold=True, fill=NAVY, color="FFFFFF")
put(ws, 17, 1, "Résultat net cumulé", align="left", bold=True, fill=BLUE, color="FFFFFF")
put(ws, 17, 2, "=B16", fmt=GNF, align="right", bold=True, fill=BLUE, color="FFFFFF")
for ci in mcols[1:]:
    put(ws, 17, ci, f"={L(ci-1)}17+{L(ci)}16", fmt=GNF, align="right", bold=True, fill=BLUE, color="FFFFFF")
put(ws, 17, 8, "=G17", fmt=GNF, align="right", bold=True, fill=BLUE, color="FFFFFF")
pnl_net_cum = "'Resultat 6 mois'!$G$17"
pnl_ca_total = "'Resultat 6 mois'!$H$9"

# =====================================================================
# 10. SYNTHESE INVESTISSEUR
# =====================================================================
ws = wb.create_sheet("Synthese Investisseur")
ws.sheet_view.showGridLines = False
for col, w in [("A", 48), ("B", 20), ("C", 14), ("D", 36)]:
    ws.column_dimensions[col].width = w
title_bar(ws, "SYNTHÈSE POUR L'INVESTISSEUR")
MS = f"'Mise en place site'!$B${fixe_mensuel_row}"
CAPEX = f"'Mise en place site'!$B${capex_total_row}"
STOCK = f"Approvisionnement!$B${appro_stock_row}"

section(ws, 3, "RENTABILITÉ UNITAIRE & SEUIL DE RENTABILITÉ (BREAK-EVEN)", ncol=4)
header_row(ws, 4, ["Indicateur", "Valeur", "Unité", "Note"])
def srow(r, lab, val, unit, note, fmt=None, bold=False, fill=None, color=None):
    put(ws, r, 1, lab, align="left", bold=bold, fill=fill, color=color)
    put(ws, r, 2, val, fmt=fmt, align="right", bold=bold, fill=fill, color=color)
    put(ws, r, 3, unit, align="center", fill=fill, color=color)
    put(ws, r, 4, note, note=(not bold), align="left", fill=fill, color=color)
srow(5, "Marge contributive / commande livrée",
     f"={H(R['price_gnf'])}-{H(R['landed_gnf'])}-{H(R['rapido'])}", "GNF",
     "Prix – coût rendu (fret inclus) – Rapido", fmt=GNF)
srow(6, "Marge nette / commande livrée", f"={couts_marge_nette_cell}", "GNF", "Après pub allouée", fmt=GNF)
srow(7, "Charges fixes journalières (pub + site)",
     f"={H(R['pub_gnf'])}+{MS}/{H(R['jours'])}", "GNF/j", "À couvrir chaque jour", fmt=GNF)
srow(8, "Break-even — commandes livrées / jour", "=B7/B5", "cmd/j",
     "Charges fixes ÷ marge contributive", fmt='#,##0.0')
srow(9, "Break-even — commandes confirmées / jour", f"=B8/{H(R['taux'])}", "cmd/j",
     "Livrées ÷ taux de livraison", fmt='#,##0.0')
srow(10, "Objectif confirmées / jour (plein régime)", f"={H(R['obj'])}", "cmd/j", "Pour comparaison", fmt=NUM)
srow(11, "Marge de sécurité vs objectif", f"={H(R['obj'])}/B9-1", "%",
     "Au-dessus du seuil", fmt=PCT, bold=True, fill=GREYBG)

section(ws, 13, "PLAN DE FINANCEMENT — UTILISATION DES FONDS", ncol=4)
header_row(ws, 14, ["Poste", "Montant (GNF)", "% du total", "Note"])
fin = [
    ("CAPEX — mise en place du site", f"={CAPEX}", "Investissement initial site"),
    ("Stock initial (stock de sécurité, fret inclus)", f"={STOCK}", "Approvisionnement"),
    ("Budget publicité (3 mois)", f"='Budget Pub Meta'!$B$8*3", "Amorçage acquisition"),
    ("Fonds de roulement (1 mois charges variables)",
     f"=({H(R['livr'])}*{H(R['jours'])})*({H(R['landed_gnf'])}+{H(R['rapido'])})",
     "Coût rendu + Rapido, 1 mois plein régime"),
]
fin_start = 15
total_fin_row = fin_start + len(fin)
r = fin_start
for lab, val, note in fin:
    put(ws, r, 1, lab, align="left")
    put(ws, r, 2, val, fmt=GNF, align="right")
    put(ws, r, 3, f"=B{r}/$B${total_fin_row}", fmt=PCT, align="right")
    put(ws, r, 4, note, note=True, align="left")
    r += 1
put(ws, r, 1, "FINANCEMENT TOTAL DEMANDÉ", bold=True, align="left", fill=NAVY, color="FFFFFF")
put(ws, r, 2, f"=SUM(B{fin_start}:B{r-1})", fmt=GNF, bold=True, align="right", fill=NAVY, color="FFFFFF")
put(ws, r, 3, f"=SUM(C{fin_start}:C{r-1})", fmt=PCT, bold=True, align="right", fill=NAVY, color="FFFFFF")
put(ws, r, 4, "", fill=NAVY)

# Financement en USD (repère)
r += 1
put(ws, r, 1, "Financement total (équivalent USD)", align="left")
put(ws, r, 2, f"=B{total_fin_row}/{H(R['usd'])}", fmt=USD, align="right")
put(ws, r, 3, "", border=True)
put(ws, r, 4, "Repère en dollars", note=True, align="left")

rr = r + 2
section(ws, rr, "RETOUR SUR INVESTISSEMENT (6 MOIS)", ncol=4); rr += 1
header_row(ws, rr, ["Indicateur", "Valeur", "Unité", "Note"]); rr += 1
srow(rr, "Résultat net cumulé sur 6 mois", f"={pnl_net_cum}", "GNF", "Onglet Résultat 6 mois", fmt=GNF); rr += 1
srow(rr, "Chiffre d'affaires cumulé 6 mois", f"={pnl_ca_total}", "GNF", "", fmt=GNF); rr += 1
srow(rr, "ROI sur 6 mois", f"={pnl_net_cum}/B{total_fin_row}", "%",
     "Résultat cumulé ÷ financement", fmt=PCT, bold=True, fill=GREYBG); rr += 1
srow(rr, "Marge nette moyenne", f"={pnl_net_cum}/{pnl_ca_total}", "%", "Résultat ÷ CA", fmt=PCT); rr += 1

rr += 1
section(ws, rr, "PRINCIPAUX FACTEURS DE RISQUE", ncol=4); rr += 1
risks = [
    "Taux de change USD/GNF : une dépréciation du GNF renchérit les achats et le fret — sensibilité testable via l'onglet Hypothèses.",
    "Taux de livraison Rapido (COD) : chaque point de taux de livraison perdu réduit directement le CA et gaspille du budget pub.",
    "Coût du fret / transport : tarif au poids/volume à confirmer avec le transitaire (base actuelle ≈ 2,65 $/article).",
    "Coût publicitaire Meta : le CPA peut augmenter ; le budget pub (20–40 $/j) est la principale charge variable à piloter.",
]
for risk in risks:
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
    c = ws.cell(rr, 1, "•  " + risk)
    c.font = Font(name=FONT, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[rr].height = 28
    rr += 1

out = "/Users/mac/HPSHOP/Business_Plan_HPSHOP_Ecommerce.xlsx"
wb.save(out)
print("Saved:", out)
